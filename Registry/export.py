import os
import sys
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError

import DBManager

class ExportHistory():
    def __init__(self, opt, date_from, date_to):
        self.table = opt

        db = DBManager.DBManager('./test.db')

        if self.table == 'Hive':
            self.hive_history = db.select_history('Hive', date_from, date_to)
            self.hiveExport2Excel(self.hive_history)

        elif self.table == 'GeneralFile':
            self.general_history = db.select_history('GeneralFile', date_from, date_to)
            self.generalExport2Excel(self.general_history)

        elif self.table == 'urls':
            self.chrome_history, self.whale_history = db.select_history('urls', date_from, date_to)
            self.urlExport2Excel(self.chrome_history, self.whale_history)

    def hiveExport2Excel(self, hive):
        wb = Workbook()
        ws = wb.active

        col_name = ['파일 타입', 'Key', 'Value Type', 'Value Name', 'Value', '수정시간']
        for i in range(1, 7):
            ws.cell(row=1, column=i).value = col_name[i - 1]

        for i, val_i in enumerate(hive):
            for j, val_j in enumerate(val_i):

                try:
                    ws.cell(row=i + 2, column=j + 1).value = hive[i][j].encode('utf-8').strip()
                except:
                    IllegalCharacterError()

        wb.save(filename='./registry history.xlsx')
        os.system('start excel.exe "%s\\registry history.xlsx"' % (sys.path[0],))

    def generalExport2Excel(self, general):
        wb = Workbook()
        ws = wb.active

        col_name = ['파일 이름', '확장자', '시그니처', '크기(byte)', '생성시간', '수정시간', '접근시간']
        for i in range(1, 8):
            ws.cell(row=1, column=i).value = col_name[i-1]

        for i, val_i in enumerate(general):
            for j, val_j in enumerate(val_i):
                ws.cell(row=i + 2, column=j + 1).value = general[i][j]

        wb.save(filename='./general file history.xlsx')
        os.system('start excel.exe "%s\\general history.xlsx"' % (sys.path[0],))

    def urlExport2Excel(self, chrome, whale):
        wb = Workbook()

        ws1 = wb.active
        ws1.title = 'chrome'
        ws2 = wb.create_sheet('whale')

        col_name = ['URL', '제목', '접근시간']
        for i in range(1, 4):
            ws1.cell(row=1, column=i).value = col_name[i - 1]
            ws2.cell(row=1, column=i).value = col_name[i - 1]

        for i, val_i in enumerate(chrome):
            for j, val_j in enumerate(val_i):
                ws1.cell(row=i + 2, column=j + 1).value = chrome[i][j].encode('utf-8').strip()


        for i, val_i in enumerate(whale):
            for j, val_j in enumerate(val_i):
                ws2.cell(row=i + 2, column=j + 1).value = whale[i][j].encode('utf-8').strip()

        wb.save(filename='./url history.xlsx')
        os.system('start excel.exe "%s\\url history.xlsx"' % (sys.path[0],))


if __name__ == '__main__':
    myHistory = ExportHistory('urls', '2020-05-01', '2020-06-31')
